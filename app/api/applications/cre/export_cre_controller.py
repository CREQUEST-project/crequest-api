import io
from fastapi import HTTPException, status
from sqlmodel import Session, select

from api.applications.cre.search_cre_controller import find_sequence_in_database, rev_comp_st
from models.factors_function_labels import FactorsFunctionLabels
from utils import send_email_attach_file_stream
from models.factors import CreResultSendEmail, Factors, MotifSearch, MotifSearchOut, Position

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from core.config import settings

from models.base import Message


def export_excel(session: Session, data_in: list[MotifSearch]):
    workbook = Workbook()

    # Iterate over each sequence and aggregate results
    all_forward_matches = []
    all_reverse_matches = []

    for motif_search in data_in:
        data_matches = _search_for_cre(session, motif_search)
        for match in data_matches["forward_strand_matches"]:
            match["original_sequence"] = motif_search.sequence
        for match in data_matches["reverse_strand_matches"]:
            match["original_sequence"] = motif_search.sequence
        all_forward_matches.extend(data_matches["forward_strand_matches"])
        all_reverse_matches.extend(data_matches["reverse_strand_matches"])

    # Forward sheet
    forward_sheet = workbook.active
    forward_sheet.title = "Forward Strand"
    _add_header_row(forward_sheet)
    _add_factor_rows(forward_sheet, all_forward_matches)

    # Reverse sheet
    reverse_sheet = workbook.create_sheet(title="Reverse Strand")
    _add_header_row(reverse_sheet)
    _add_factor_rows(reverse_sheet, all_reverse_matches)

    # Format sheets
    _format_sheet(forward_sheet)
    _format_sheet(reverse_sheet)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return output


def _add_header_row(sheet):
    sheet.append(["original_sequence", "ac", "dt", "de", "kw", "os", "ra", "rt", "rl", "rd", "sq", "positions"])

    # Style header row
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=14, bold=True)


def _add_factor_rows(sheet, matches):
    for match in matches:
        positions = "; ".join([f"{pos.start}-{pos.end}" for pos in match["positions"]])
        sheet.append(
            [
                match["original_sequence"],  # Add original sequence
                match["factor_id"],
                match["dt"], 
                match["de"],
                match["kw"],
                match["os"],
                match["ra"],
                match["rt"],
                match["rl"],
                match["rd"],
                match["sq"],
                positions,
            ]
        )


def _format_sheet(sheet):
    # Align center for all cells
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto size columns
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width


def send_cre_excel_email(session: Session, data_in: list[CreResultSendEmail]):
    output = export_excel(session, data_in)

    # Send email with attachment
    send_email_attach_file_stream(
        receiver_email=data_in.receiver_email,
        subject="Excel Export",
        body="Please find attached the excel file.",
        sender_email=settings.EMAIL_HOST_USER,
        file_stream=output,
    )

    return Message(status_code=status.HTTP_200_OK, message="Email sent successfully.")


def _search_for_cre(session: Session, data_in: MotifSearch) -> MotifSearchOut:
    reverse_complement = rev_comp_st(data_in.sequence)
    db_factors = session.exec(select(Factors).order_by(Factors.ft_id)).all()
    database = {factor.ac: factor.sq for factor in db_factors}

    # Find matches on both strands
    forward_matches = find_sequence_in_database(data_in.sequence, database)
    reverse_matches = find_sequence_in_database(reverse_complement, database)

    # Group matches by factor_id
    forward_matches_grouped = {}
    for match in forward_matches:
        factor_id = match[1]
        if factor_id not in forward_matches_grouped:
            forward_matches_grouped[factor_id] = []
        forward_matches_grouped[factor_id].append(
            Position(start=match[2], end=match[3])
        )  # Store start, end

    reverse_matches_grouped = {}
    for match in reverse_matches:
        factor_id = match[1]
        if factor_id not in reverse_matches_grouped:
            reverse_matches_grouped[factor_id] = []
        reverse_matches_grouped[factor_id].append(
            Position(start=match[2], end=match[3])
        )  # Store start, end

    # Prepare data for matches with color, only one entry per factor_id
    forward_matches_with_color = []
    for factor_id, positions in forward_matches_grouped.items():
        factor = session.exec(select(Factors).where(Factors.ac == factor_id)).first()
        if not factor:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Factor with AC {factor_id} not found.",
            )

        function_label = None
        if factor.ft_id:
            function_label = session.exec(
                select(FactorsFunctionLabels).where(
                    FactorsFunctionLabels.id == factor.ft_id
                )
            ).first()

        forward_matches_with_color.append(
            {
                "factor_id": factor_id,
                "sq": factor.sq,
                "de": factor.de,
                "dt": factor.dt,  # Extract dt
                "kw": factor.kw,  # Extract kw
                "os": factor.os,  # Extract os
                "ra": factor.ra,  # Extract ra
                "rt": factor.rt,  # Extract rt
                "rl": factor.rl,  # Extract rl
                "rd": factor.rd,  # Extract rd
                "function_label": function_label,
                "positions": positions,  # Store start, end as an array
                "color": factor.color,
            }
        )

    reverse_matches_with_color = []
    for factor_id, positions in reverse_matches_grouped.items():
        factor = session.exec(select(Factors).where(Factors.ac == factor_id)).first()
        if not factor:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Factor with AC {factor_id} not found.",
            )
        function_label = None
        if factor.ft_id:
            function_label = session.exec(
                select(FactorsFunctionLabels).where(
                    FactorsFunctionLabels.id == factor.ft_id
                )
            ).first()

        reverse_matches_with_color.append(
            {
                "factor_id": factor_id,
                "sq": factor.sq,
                "de": factor.de,
                "dt": factor.dt,  # Extract dt
                "kw": factor.kw,  # Extract kw
                "os": factor.os,  # Extract os
                "ra": factor.ra,  # Extract ra
                "rt": factor.rt,  # Extract rt
                "rl": factor.rl,  # Extract rl
                "rd": factor.rd,  # Extract rd
                "function_label": function_label,
                "positions": positions,  # Store start, end as an array
                "color": factor.color,
            }
        )

    data = {
        "original_sequence": data_in.sequence,
        "reverse_complement_sequence": reverse_complement,
        "forward_strand_matches": forward_matches_with_color,
        "reverse_strand_matches": reverse_matches_with_color,
    }
    return data